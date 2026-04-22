from datetime import date
import csv
import io
import json
import os
import re
import shutil
import subprocess
import tempfile
from typing import Optional
from openpyxl import load_workbook

from app.models import (
    POINT_RULE,
    Database,
    SportsRepository,
    calc_age_group,
    scoring_strategy_for_event_type,
)


class SportsMeetService:
    CLEAR_TABLES = {
        "settings": "系统设置",
        "departments": "部门",
        "events": "项目",
        "competitive_athletes": "竞技运动员",
        "fun_athletes": "趣味运动员",
        "teams": "队伍",
        "team_members": "队伍成员",
        "athlete_registrations": "报名记录",
        "results": "成绩记录",
    }
    REPORT_ENV_KEYS = [
        "date",
        "wind_direction",
        "wind_speed",
        "air_quality",
        "weather",
        "temperature_high",
        "temperature_low",
    ]

    def __init__(self, db_path: str) -> None:
        self.db = Database(db_path)

    def init_db(self) -> None:
        self.db.initialize()

    def set_meet_date(self, meet_date: date) -> None:
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            repo.set_meet_date(meet_date.isoformat())
            conn.commit()

    def set_report_environment_settings(self, payload: dict[str, str]) -> None:
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            for key in self.REPORT_ENV_KEYS:
                val = str(payload.get(key, "")).strip()
                repo.set_setting(f"report_env.{key}", val)
            conn.commit()

    def get_report_environment_settings(self) -> dict[str, str]:
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            return {
                key: str(repo.get_setting(f"report_env.{key}") or "").strip()
                for key in self.REPORT_ENV_KEYS
            }

    def list_notice_templates(self, template_dir: str) -> list[str]:
        if not os.path.isdir(template_dir):
            return []
        names: list[str] = []
        for name in os.listdir(template_dir):
            lower = name.lower()
            if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
                names.append(name)
        names.sort()
        return names

    def _event_gender_label(self, gender: str) -> str:
        if gender == "male":
            return "男子"
        if gender == "female":
            return "女子"
        return "混合"

    def _event_group_label(self, age_group: str) -> str:
        if age_group == "A":
            return "甲组"
        if age_group == "B":
            return "乙组"
        if age_group == "C":
            return "丙组"
        return "不限组"

    def _event_display_name(self, event: dict) -> str:
        return f"{event.get('name', '')}{self._event_gender_label(str(event.get('gender', '')))}{self._event_group_label(str(event.get('age_group', '')))}"

    def _notice_title_for_event(self, event: dict, layout: Optional[dict] = None) -> str:
        title_map = (layout or {}).get("notice_title_by_event_type", {}) if isinstance(layout, dict) else {}
        event_type = str(event.get("event_type", "")).strip()
        if isinstance(title_map, dict):
            custom = str(title_map.get(event_type, "")).strip()
            if custom:
                return custom
            custom_default = str(title_map.get("default", "")).strip()
            if custom_default:
                return custom_default
        if event_type == "track":
            return "径赛成绩单"
        if event_type == "field":
            return "田赛成绩单"
        return "趣味成绩单"

    def get_meet_date(self) -> date:
        with self.db.connect() as conn:
            iso = SportsRepository(conn).get_meet_date_iso()
            return date.fromisoformat(iso) if iso else date(2026, 4, 23)

    def calc_age_group(self, gender: str, birth_date: date) -> str:
        return calc_age_group(gender, birth_date, self.get_meet_date())

    def _validate_athlete_type(self, athlete_type: str) -> str:
        if athlete_type not in {"competitive", "fun"}:
            raise ValueError("athlete_type 必须为 competitive 或 fun")
        return athlete_type

    def _expand_event_genders(self, gender_raw: str) -> list[str]:
        gender = (gender_raw or "").strip().lower()
        aliases = {
            "male": ["male"],
            "female": ["female"],
            "mixed": ["mixed"],
            "男女": ["male", "female"],
            "both": ["male", "female"],
            "all": ["male", "female"],
            "male+female": ["male", "female"],
            "male/female": ["male", "female"],
            "mf": ["male", "female"],
        }
        if gender in aliases:
            return aliases[gender]
        raise ValueError("gender 必须为 male/female/mixed，或男女合并值（男女/both/male+female/all）")

    def _parse_performance_numeric(self, strategy: str, performance: Optional[str]) -> Optional[float]:
        text = self._normalize_performance_text(strategy, performance)
        if not text:
            return None
        if strategy == "time":
            if not re.fullmatch(r"\d+(?:\.\d+)*", text):
                return None
            parts = text.split(".")
            if len(parts) <= 2:
                return float(text)
            minute = int(parts[0])
            sec_int = parts[1]
            sec_decimal = "".join(parts[2:])
            sec_val = float(sec_int + ("." + sec_decimal if sec_decimal else ""))
            return minute * 60 + sec_val
        if strategy == "length":
            if not re.fullmatch(r"\d+(?:\.\d+)*", text):
                return None
            parts = text.split(".")
            if len(parts) <= 2:
                return float(text)
            return float(parts[0] + "." + "".join(parts[1:]))
        if strategy == "count":
            if not re.fullmatch(r"\d+", text):
                return None
            return float(int(text))
        return None

    def _normalize_performance_text(self, strategy: str, performance: Optional[str]) -> Optional[str]:
        raw = (performance or "").strip()
        if not raw:
            return None
        text = raw.replace("：", ":").replace("．", ".").replace("。", ".")
        text = re.sub(r"\s+", "", text)

        if strategy == "time":
            text = text.lower().replace("分", ".").replace("秒", "")
            text = text.replace(":", ".")
            text = text.replace("s", "")
        elif strategy == "length":
            text = text.lower().replace("米", "").replace("m", "")
            text = text.replace(":", ".")
        elif strategy == "count":
            text = text.replace("次", "").replace("个", "")
            text = text.replace(":", "")

        if strategy == "count":
            text = re.sub(r"[^0-9]", "", text)
            text = text.strip()
        else:
            text = re.sub(r"[^0-9.]", "", text)
            text = text.strip(".")
        return text or None

    def _format_time_seconds(self, seconds: float) -> str:
        total_cs = max(int(round(seconds * 100)), 0)
        minute = total_cs // 6000
        sec_cs = total_cs % 6000
        sec_int = sec_cs // 100
        sec_frac = sec_cs % 100
        return f"{minute:02d}:{sec_int:02d}.{sec_frac:02d}"

    def _format_performance_for_display(self, scoring_strategy: str, performance: Optional[str]) -> str:
        text = (performance or "").strip()
        if not text:
            return ""
        if scoring_strategy == "time":
            sec = self._parse_performance_numeric("time", text)
            if sec is None:
                return text
            return self._format_time_seconds(sec)
        if scoring_strategy == "length":
            meters = self._parse_performance_numeric("length", text)
            if meters is None:
                return text
            return f"{meters:.2f}m"
        if scoring_strategy == "count":
            count_val = self._parse_performance_numeric("count", text)
            if count_val is None:
                return text
            return f"{int(count_val)}个"
        return text

    def _format_result_rows_performance(self, rows: list[dict]) -> list[dict]:
        for row in rows:
            strategy = str(row.get("scoring_strategy", "")).strip()
            if strategy in {"time", "length", "count"}:
                row["performance"] = self._format_performance_for_display(strategy, row.get("performance"))
        return rows

    def _auto_rank_for_result(
        self,
        repo: SportsRepository,
        event_id: int,
        scoring_strategy: str,
        performance: Optional[str],
    ) -> int:
        rows = repo.list_event_results(event_id)
        if not rows:
            return 1
        max_rank = max(int(r["rank"]) for r in rows)
        new_val = self._parse_performance_numeric(scoring_strategy, performance)
        if new_val is None:
            return max_rank + 1

        better = 0
        comparable = 0
        for row in rows:
            old_val = self._parse_performance_numeric(scoring_strategy, row["performance"])
            if old_val is None:
                continue
            comparable += 1
            if scoring_strategy == "time":
                if old_val < new_val:
                    better += 1
            else:
                if old_val > new_val:
                    better += 1
        if comparable == 0:
            return max_rank + 1
        return better + 1

    def _recalculate_event_ranks(self, repo: SportsRepository, event_id: int, scoring_strategy: str) -> None:
        rows = [dict(r) for r in repo.list_event_results(event_id)]
        if not rows:
            return

        def _sort_key(item: dict):
            val = self._parse_performance_numeric(scoring_strategy, item.get("performance"))
            if val is None:
                return (1, float("inf"), int(item["id"]))
            if scoring_strategy == "time":
                return (0, val, int(item["id"]))
            return (0, -val, int(item["id"]))

        ranked = sorted(rows, key=_sort_key)
        for idx, row in enumerate(ranked, start=1):
            repo.update_result_rank_points(int(row["id"]), idx, POINT_RULE.get(idx, 0))

    def add_department(self, name: str, total_members: int) -> int:
        with self.db.connect() as conn:
            department_id = SportsRepository(conn).insert_department(name, total_members)
            conn.commit()
            return department_id

    def add_athlete(
        self,
        name: str,
        gender: str,
        department_id: int,
        athlete_no: Optional[str] = None,
        age_group: Optional[str] = None,
        athlete_type: str = "competitive",
    ) -> int:
        athlete_type = self._validate_athlete_type(athlete_type)
        if gender not in ("male", "female"):
            raise ValueError("gender 必须是 male 或 female")
        if age_group is not None and age_group != "" and age_group not in ("A", "B", "C"):
            raise ValueError("age_group 必须是 A/B/C")
        resolved_group = age_group if age_group else None

        with self.db.connect() as conn:
            athlete_id = SportsRepository(conn).insert_athlete(
                athlete_type=athlete_type,
                athlete_no=athlete_no,
                name=name,
                gender=gender,
                department_id=department_id,
                age_group=resolved_group,
            )
            conn.commit()
            return athlete_id

    def list_events(self):
        with self.db.connect() as conn:
            return SportsRepository(conn).list_events()

    def list_athletes(self):
        with self.db.connect() as conn:
            return SportsRepository(conn).list_athletes_with_department()

    def list_registration_pairs(self):
        with self.db.connect() as conn:
            return SportsRepository(conn).list_registration_pairs()

    def query_athletes(self, athlete_type: str = "", keyword: str = "") -> list[dict]:
        type_filter = (athlete_type or "").strip()
        if type_filter and type_filter not in {"competitive", "fun"}:
            raise ValueError("athlete_type 必须为 competitive/fun 或留空")
        kw = (keyword or "").strip().lower()

        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            athletes = [dict(r) for r in repo.list_athletes_with_department()]
            pairs = [dict(r) for r in repo.list_registration_pairs()]
            events = [dict(r) for r in repo.list_events()]

        event_map = {int(e["id"]): e for e in events}
        reg_map: dict[tuple[str, int], list[int]] = {}
        for p in pairs:
            key = (str(p["athlete_type"]), int(p["athlete_ref_id"]))
            reg_map.setdefault(key, []).append(int(p["event_id"]))

        def _event_label(event: dict) -> str:
            g = self._event_gender_label(str(event.get("gender", "")))
            ag = self._event_group_label(str(event.get("age_group", "")))
            return f"{event.get('name', '')}{g}{ag}"

        items: list[dict] = []
        for a in athletes:
            at = str(a.get("athlete_type", ""))
            if type_filter and at != type_filter:
                continue

            hay = " ".join(
                [
                    str(a.get("athlete_no", "") or ""),
                    str(a.get("name", "") or ""),
                    str(a.get("department_name", "") or ""),
                ]
            ).lower()
            if kw and kw not in hay:
                continue

            key = (at, int(a.get("athlete_ref_id", 0)))
            event_ids = reg_map.get(key, [])
            labels = []
            for eid in event_ids:
                e = event_map.get(eid)
                if e and int(e.get("is_individual", 0)) == 1:
                    labels.append(_event_label(e))
            labels = sorted(set(labels))

            items.append(
                {
                    "athlete_type": at,
                    "athlete_ref_id": a.get("athlete_ref_id"),
                    "athlete_no": a.get("athlete_no", "") or "",
                    "name": a.get("name", "") or "",
                    "gender": a.get("gender", "") or "",
                    "age_group": a.get("age_group", "") or "",
                    "department_name": a.get("department_name", "") or "",
                    "registration_count": len(labels),
                    "registered_events": "；".join(labels),
                }
            )
        items.sort(key=lambda x: (str(x["athlete_type"]), str(x["athlete_no"]), str(x["name"])))
        return items

    def get_registered_individual_events(self, athlete_type: str, athlete_no: str) -> list[dict]:
        athlete_type = self._validate_athlete_type((athlete_type or "").strip())
        athlete_no_text = (athlete_no or "").strip()
        if not athlete_no_text:
            raise ValueError("athlete_no 不能为空")

        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            athlete = repo.get_athlete_by_no(athlete_type, athlete_no_text)
            if not athlete:
                raise ValueError(f"运动员不存在: {athlete_type}/{athlete_no_text}")
            rows = repo.list_registered_individual_events_for_athlete(
                athlete_type=athlete_type,
                athlete_ref_id=int(athlete["athlete_ref_id"]),
            )
            items = [dict(r) for r in rows]

        for it in items:
            it["label"] = (
                f"{it.get('name', '')}"
                f"{self._event_gender_label(str(it.get('gender', '')))}"
                f"{self._event_group_label(str(it.get('age_group', '')))}"
            )
        return items

    def list_individual_events_by_category(self, category: str):
        if category not in {"competitive", "fun"}:
            raise ValueError("category 必须为 competitive 或 fun")
        with self.db.connect() as conn:
            return SportsRepository(conn).list_individual_events_by_category(category)

    def add_athlete_by_department_name(
        self,
        athlete_type: str,
        athlete_no: Optional[str],
        name: str,
        gender: str,
        department_name: str,
        age_group: Optional[str] = None,
    ) -> int:
        athlete_type = self._validate_athlete_type(athlete_type)
        athlete_no_text = (athlete_no or "").strip()
        if not athlete_no_text:
            raise ValueError("athlete_no 不能为空")
        name_text = (name or "").strip()
        if not name_text:
            raise ValueError("name 不能为空")
        if gender not in {"male", "female"}:
            raise ValueError("gender 必须是 male 或 female")
        dept_name = (department_name or "").strip()
        if not dept_name:
            raise ValueError("department_name 不能为空")
        age_group_text = (age_group or "").strip() or None
        if age_group_text and age_group_text not in {"A", "B", "C"}:
            raise ValueError("age_group 必须是 A/B/C")

        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            exists = repo.get_athlete_by_no(athlete_type, athlete_no_text)
            if exists:
                raise ValueError(f"athlete_no 已存在: {athlete_no_text}")
            dept = repo.get_department_by_name(dept_name)
            if dept:
                dept_id = int(dept["id"])
            else:
                dept_id = repo.insert_department(dept_name, 0)
            athlete_id = repo.insert_athlete(
                athlete_type=athlete_type,
                athlete_no=athlete_no_text,
                name=name_text,
                gender=gender,
                department_id=dept_id,
                age_group=age_group_text,
            )
            conn.commit()
            return athlete_id

    def delete_athlete_by_no(self, athlete_type: str, athlete_no: str) -> dict:
        athlete_type = self._validate_athlete_type(athlete_type)
        athlete_no_text = (athlete_no or "").strip()
        if not athlete_no_text:
            raise ValueError("athlete_no 不能为空")
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            athlete = repo.get_athlete_by_no(athlete_type, athlete_no_text)
            if not athlete:
                raise ValueError(f"运动员不存在: {athlete_type}/{athlete_no_text}")
            athlete_ref_id = int(athlete["athlete_ref_id"])
            related = repo.delete_athlete_related_data(athlete_type, athlete_ref_id)
            deleted = repo.delete_athlete_by_id(athlete_type, athlete_ref_id)
            conn.commit()
            return {
                "deleted_athlete": deleted,
                "deleted_related": related,
                "athlete_type": athlete_type,
                "athlete_no": athlete_no_text,
            }

    def adjust_athlete_registration(self, athlete_type: str, athlete_no: str, event_id: int, op: str) -> dict:
        athlete_type = self._validate_athlete_type(athlete_type)
        athlete_no_text = (athlete_no or "").strip()
        if not athlete_no_text:
            raise ValueError("athlete_no 不能为空")
        if op not in {"add", "remove"}:
            raise ValueError("op 必须是 add 或 remove")

        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            athlete = repo.get_athlete_by_no(athlete_type, athlete_no_text)
            if not athlete:
                raise ValueError(f"运动员不存在: {athlete_type}/{athlete_no_text}")
            athlete_ref_id = int(athlete["athlete_ref_id"])
            event = repo.get_event_by_id(event_id)
            if not event:
                raise ValueError(f"比赛项目不存在: {event_id}")
            if event["category"] != athlete_type:
                raise ValueError("项目类别与运动员类型不匹配")
            if int(event["is_individual"]) != 1:
                raise ValueError("仅个人项目支持运动员报名")
            if event["gender"] not in {athlete["gender"], "mixed"}:
                raise ValueError("项目性别与运动员不匹配")

            exists = repo.athlete_registration_exists(athlete_type, athlete_ref_id, event_id)
            changed = 0
            if op == "add":
                if exists:
                    conn.commit()
                    return {"changed": 0, "status": "exists"}
                repo.insert_athlete_registration(athlete_type, athlete_ref_id, event_id)
                changed = 1
            else:
                changed = repo.delete_athlete_registration(athlete_type, athlete_ref_id, event_id)

            conn.commit()
            return {
                "changed": changed,
                "status": "ok",
                "athlete_type": athlete_type,
                "athlete_no": athlete_no_text,
                "event_id": event_id,
                "op": op,
            }

    def register_athlete_event(self, athlete_type: str, athlete_ref_id: int, event_id: int) -> int:
        athlete_type = self._validate_athlete_type(athlete_type)
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            athlete = repo.get_athlete_by_id(athlete_type, athlete_ref_id)
            if not athlete:
                raise ValueError(f"运动员不存在: {athlete_type}/{athlete_ref_id}")
            event = repo.get_event_by_id(event_id)
            if not event:
                raise ValueError(f"比赛项目不存在: {event_id}")

            if event["category"] != athlete_type:
                raise ValueError("项目类别与运动员类型不匹配")
            if event["is_individual"] != 1:
                raise ValueError("该项目为集体项目，请使用组队流程")
            if event["gender"] != athlete["gender"]:
                raise ValueError("项目性别与运动员不匹配")
            reg_id = repo.insert_athlete_registration(athlete_type, athlete_ref_id, event_id)
            conn.commit()
            return reg_id

    def create_team(self, department_id: int, event_id: int, name: str) -> int:
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            event = repo.get_event_by_id(event_id)
            if not event:
                raise ValueError(f"比赛项目不存在: {event_id}")
            if event["is_individual"] != 0:
                raise ValueError("仅集体项目可以创建队伍")
            team_id = repo.insert_team(department_id, event_id, name)
            conn.commit()
            return team_id

    def add_team_member(self, team_id: int, athlete_type: str, athlete_ref_id: int) -> int:
        athlete_type = self._validate_athlete_type(athlete_type)
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            team = repo.get_team_by_id(team_id)
            if not team:
                raise ValueError(f"队伍不存在: {team_id}")
            athlete = repo.get_athlete_by_id(athlete_type, athlete_ref_id)
            if not athlete:
                raise ValueError(f"运动员不存在: {athlete_type}/{athlete_ref_id}")
            if athlete["department_id"] != team["department_id"]:
                raise ValueError("队伍成员必须来自同一部门")

            event = repo.get_event_by_id(int(team["event_id"]))
            if event and event["category"] != athlete_type:
                raise ValueError("队伍项目类别与运动员类型不匹配")

            team_member_id = repo.insert_team_member(team_id, athlete_type, athlete_ref_id)
            conn.commit()
            return team_member_id

    def list_team_events(self) -> list[dict]:
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            items = [dict(r) for r in repo.list_events()]
            return [i for i in items if int(i.get("is_individual", 1)) == 0]

    def query_teams(
        self,
        keyword: str = "",
        department_name: str = "",
        event_id: Optional[int] = None,
    ) -> list[dict]:
        kw = (keyword or "").strip().lower()
        dept = (department_name or "").strip()
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            teams = [dict(r) for r in repo.list_teams_with_details()]

            items: list[dict] = []
            for t in teams:
                if dept and str(t.get("department_name", "")) != dept:
                    continue
                if event_id is not None and int(t.get("event_id", 0)) != int(event_id):
                    continue
                hay = " ".join(
                    [
                        str(t.get("team_name", "") or ""),
                        str(t.get("department_name", "") or ""),
                        str(t.get("event_name", "") or ""),
                    ]
                ).lower()
                if kw and kw not in hay:
                    continue

                members = [dict(r) for r in repo.list_team_members_with_details(int(t["id"]))]
                member_names = []
                for m in members:
                    at = str(m.get("athlete_type", ""))
                    prefix = "竞" if at == "competitive" else ("趣" if at == "fun" else "")
                    member_names.append(f"{prefix}{m.get('athlete_no', '')}/{m.get('athlete_name', '')}")

                item = dict(t)
                item["member_count"] = len(members)
                item["members_summary"] = "；".join(member_names)
                items.append(item)

            items.sort(key=lambda x: int(x.get("id", 0)))
            return items

    def add_team_by_department_name(self, department_name: str, event_id: int, team_name: str) -> int:
        dept_name = (department_name or "").strip()
        name = (team_name or "").strip()
        if not dept_name:
            raise ValueError("department_name 不能为空")
        if not name:
            raise ValueError("team_name 不能为空")

        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            event = repo.get_event_by_id(event_id)
            if not event:
                raise ValueError(f"比赛项目不存在: {event_id}")
            if int(event["is_individual"]) != 0:
                raise ValueError("仅团体项目可建队")

            dept = repo.get_department_by_name(dept_name)
            if not dept:
                dept_id = repo.insert_department(dept_name, 0)
            else:
                dept_id = int(dept["id"])

            if repo.team_exists(dept_id, event_id, name):
                raise ValueError("该单位在此项目下已存在同名队伍")

            team_id = repo.insert_team(dept_id, event_id, name)
            conn.commit()
            return team_id

    def delete_team(self, team_id: int) -> dict:
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            team = repo.get_team_by_id(team_id)
            if not team:
                raise ValueError(f"队伍不存在: {team_id}")
            related = repo.delete_team_related_data(team_id)
            deleted = repo.delete_team_by_id(team_id)
            conn.commit()
            return {"team_id": team_id, "deleted_team": deleted, "deleted_related": related}

    def get_team_members(self, team_id: int) -> list[dict]:
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            team = repo.get_team_by_id(team_id)
            if not team:
                raise ValueError(f"队伍不存在: {team_id}")
            rows = [dict(r) for r in repo.list_team_members_with_details(team_id)]
            for r in rows:
                at = str(r.get("athlete_type", ""))
                r["athlete_type_text"] = "竞技" if at == "competitive" else ("趣味" if at == "fun" else at)
                g = str(r.get("gender", ""))
                r["gender_text"] = "男" if g == "male" else ("女" if g == "female" else g)
            return rows

    def adjust_team_member(self, team_id: int, athlete_type: str, athlete_no: str, op: str) -> dict:
        athlete_type = self._validate_athlete_type((athlete_type or "").strip())
        athlete_no_text = (athlete_no or "").strip()
        if not athlete_no_text:
            raise ValueError("athlete_no 不能为空")
        if op not in {"add", "remove"}:
            raise ValueError("op 必须是 add 或 remove")

        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            team = repo.get_team_by_id(team_id)
            if not team:
                raise ValueError(f"队伍不存在: {team_id}")
            athlete = repo.get_athlete_by_no(athlete_type, athlete_no_text)
            if not athlete:
                raise ValueError(f"运动员不存在: {athlete_type}/{athlete_no_text}")
            athlete_ref_id = int(athlete["athlete_ref_id"])

            event = repo.get_event_by_id(int(team["event_id"]))
            if not event:
                raise ValueError("队伍关联项目不存在")
            if int(event["is_individual"]) != 0:
                raise ValueError("仅团体项目允许队伍成员操作")
            if str(event["category"]) != athlete_type:
                raise ValueError("队伍项目类别与运动员类型不匹配")
            if int(athlete["department_id"]) != int(team["department_id"]):
                raise ValueError("队伍成员必须来自同一单位")
            if str(event["gender"]) in {"male", "female"} and str(event["gender"]) != str(athlete["gender"]):
                raise ValueError("队伍项目性别与运动员不匹配")

            exists = repo.team_member_exists(team_id, athlete_type, athlete_ref_id)
            changed = 0
            status = "ok"
            if op == "add":
                if exists:
                    status = "exists"
                else:
                    repo.insert_team_member(team_id, athlete_type, athlete_ref_id)
                    changed = 1
            else:
                changed = repo.delete_team_member(team_id, athlete_type, athlete_ref_id)
                if changed == 0:
                    status = "not_found"
            conn.commit()
            return {
                "team_id": team_id,
                "athlete_type": athlete_type,
                "athlete_no": athlete_no_text,
                "op": op,
                "changed": changed,
                "status": status,
            }

    def record_result(
        self,
        event_id: int,
        rank: Optional[int] = None,
        athlete_type: Optional[str] = None,
        athlete_ref_id: Optional[int] = None,
        athlete_no: Optional[str] = None,
        team_id: Optional[int] = None,
        performance: Optional[str] = None,
    ) -> int:
        athlete_no_text = (athlete_no or "").strip()
        if athlete_ref_id is not None and athlete_no_text:
            raise ValueError("athlete_ref_id 和 athlete_no 不能同时传")
        if athlete_ref_id is None and athlete_no_text:
            if not athlete_type:
                raise ValueError("使用 athlete_no 录入时，athlete_type 必填")
            athlete_type = self._validate_athlete_type(athlete_type)
            with self.db.connect() as conn:
                repo = SportsRepository(conn)
                found = repo.get_athlete_by_no(athlete_type, athlete_no_text)
                if not found:
                    raise ValueError(f"运动员不存在: {athlete_type}/{athlete_no_text}")
                athlete_ref_id = int(found["id"])

        has_athlete = athlete_ref_id is not None
        has_team = team_id is not None
        if has_athlete == has_team:
            raise ValueError("必须且只能传 athlete_ref_id 或 team_id 其中之一")

        if has_athlete:
            if not athlete_type:
                raise ValueError("录入个人成绩时，athlete_type 必填")
            athlete_type = self._validate_athlete_type(athlete_type)

        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            event = repo.get_event_by_id(event_id)
            if not event:
                raise ValueError(f"比赛项目不存在: {event_id}")
            scoring_strategy = str(event["scoring_strategy"])
            normalized_performance = self._normalize_performance_text(scoring_strategy, performance)
            if performance and not normalized_performance:
                if scoring_strategy == "time":
                    raise ValueError("径赛成绩格式无效，请使用 '.' 分隔（示例：9.86 或 1.36.5）")
                if scoring_strategy == "length":
                    raise ValueError("田赛成绩格式无效，请使用米数数字并用 '.' 分隔（示例：6.23）")
                if scoring_strategy == "count":
                    raise ValueError("count 成绩必须为整数，单位个（示例：18）")
                raise ValueError("成绩格式无效，请使用数字并用 '.' 分隔")
            if scoring_strategy == "time":
                if normalized_performance:
                    sec = self._parse_performance_numeric("time", normalized_performance)
                    if sec is None:
                        raise ValueError("径赛成绩格式无效，请使用 '.' 分隔（示例：9.86 或 1.36.5）")
                    normalized_performance = f"{sec:.2f}"
                else:
                    normalized_performance = None
            elif scoring_strategy == "count":
                if normalized_performance:
                    cnt = self._parse_performance_numeric("count", normalized_performance)
                    if cnt is None:
                        raise ValueError("count 成绩必须为整数，单位个（示例：18）")
                    normalized_performance = str(int(cnt))
                else:
                    normalized_performance = None

            if has_athlete:
                athlete = repo.get_athlete_by_id(athlete_type or "", int(athlete_ref_id))
                if not athlete:
                    raise ValueError(f"运动员不存在: {athlete_type}/{athlete_ref_id}")
                if event["category"] != athlete_type:
                    raise ValueError("项目类别与运动员类型不匹配")
                if int(event["is_individual"]) != 1:
                    raise ValueError("该项目为团体项目，不能录入个人成绩")
                if not repo.athlete_registration_exists(athlete_type, int(athlete_ref_id), event_id):
                    raise ValueError("该运动员未报名此项目，不能录入成绩")
            if has_team:
                team = repo.get_team_by_id(int(team_id))
                if not team:
                    raise ValueError(f"队伍不存在: {team_id}")
                if int(team["event_id"]) != int(event_id):
                    raise ValueError("队伍与项目不匹配")

            auto_rank = rank is None
            final_rank = int(rank) if rank is not None else self._auto_rank_for_result(
                repo,
                event_id,
                scoring_strategy,
                normalized_performance,
            )
            if final_rank < 1:
                raise ValueError("rank 必须 >= 1")

            result_id = repo.insert_result(
                event_id=event_id,
                rank=final_rank,
                points=POINT_RULE.get(final_rank, 0),
                athlete_type=athlete_type if has_athlete else None,
                athlete_ref_id=athlete_ref_id if has_athlete else None,
                team_id=team_id if has_team else None,
                performance=normalized_performance,
            )
            if auto_rank:
                self._recalculate_event_ranks(repo, event_id, scoring_strategy)
            conn.commit()
            return result_id

    def standings(self):
        with self.db.connect() as conn:
            return SportsRepository(conn).standings()

    def participation_rate(self):
        with self.db.connect() as conn:
            return SportsRepository(conn).participation_rate()

    def dashboard_summary(self) -> dict:
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            recent_results = [dict(r) for r in repo.recent_results(10)]
            self._format_result_rows_performance(recent_results)
            return {
                "event_count": repo.events_count(),
                "dept_count": repo.departments_count(),
                "athlete_count": repo.athletes_count(),
                "standings": repo.standings()[:10],
                "recent_results": recent_results,
            }

    def workbench_data(self) -> dict:
        summary = self.dashboard_summary()
        status = self.get_initialization_status()
        alerts = []
        if not status["completed"]:
            alerts.append("系统初始化未完成，请先设置比赛日期并导入项目。")
        if summary["athlete_count"] == 0:
            alerts.append("暂无运动员数据，建议尽快导入名单。")
        return {
            **summary,
            "init_completed": status["completed"],
            "alerts": alerts,
        }

    def list_department_names(self) -> list[str]:
        with self.db.connect() as conn:
            rows = SportsRepository(conn).list_departments()
            return [row["name"] for row in rows]

    def get_data_view(self, view_name: str) -> list[dict]:
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            if view_name == "events":
                rows = repo.list_events()
            elif view_name == "athletes":
                rows = repo.list_athletes_with_department()
            elif view_name == "departments":
                rows = repo.list_departments()
            elif view_name == "teams":
                rows = repo.list_teams_with_details()
            elif view_name == "registrations":
                rows = repo.list_registrations_with_details()
            elif view_name == "results":
                rows = repo.list_results_with_details()
            elif view_name == "standings":
                rows = repo.standings()
            elif view_name == "participation":
                rows = repo.participation_rate()
            else:
                raise ValueError(f"不支持的数据视图: {view_name}")
            items = [dict(row) for row in rows]
            if view_name == "results":
                self._format_result_rows_performance(items)
            return items

    def get_grid_page(
        self,
        view_name: str,
        page: int,
        page_size: int,
        keyword: str = "",
        department_name: str = "",
        gender: str = "",
        age_group: str = "",
        category: str = "",
        scoring_strategy: str = "",
        sort_by: str = "",
        sort_dir: str = "desc",
    ) -> dict:
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            if view_name == "events":
                total, rows = repo.page_events(
                    page, page_size, keyword, gender, age_group, category, scoring_strategy, sort_by, sort_dir
                )
            elif view_name == "athletes":
                total, rows = repo.page_athletes(
                    page, page_size, keyword, department_name, gender, age_group, sort_by, sort_dir
                )
            elif view_name == "departments":
                total, rows = repo.page_departments(page, page_size, keyword, sort_by, sort_dir)
            elif view_name == "teams":
                total, rows = repo.page_teams(page, page_size, keyword, department_name, sort_by, sort_dir)
            elif view_name == "registrations":
                total, rows = repo.page_registrations(page, page_size, keyword, department_name, sort_by, sort_dir)
            elif view_name == "results":
                total, rows = repo.page_results(
                    page, page_size, keyword, department_name, category, scoring_strategy, sort_by, sort_dir
                )
            elif view_name == "standings":
                total, rows = repo.page_standings(page, page_size, sort_by, sort_dir)
            elif view_name == "participation":
                total, rows = repo.page_participation(page, page_size, sort_by, sort_dir)
            else:
                raise ValueError(f"不支持的数据视图: {view_name}")

            items = [dict(r) for r in rows]
            if view_name == "results":
                self._format_result_rows_performance(items)
            return {
                "view": view_name,
                "page": page,
                "page_size": page_size,
                "total": total,
                "pages": (total + page_size - 1) // page_size if page_size > 0 else 0,
                "sort_by": sort_by,
                "sort_dir": sort_dir,
                "columns": list(items[0].keys()) if items else [],
                "items": items,
            }

    def export_grid_csv(
        self,
        view_name: str,
        keyword: str = "",
        department_name: str = "",
        gender: str = "",
        age_group: str = "",
        category: str = "",
        scoring_strategy: str = "",
    ) -> str:
        first = self.get_grid_page(
            view_name=view_name,
            page=1,
            page_size=100000,
            keyword=keyword,
            department_name=department_name,
            gender=gender,
            age_group=age_group,
            category=category,
            scoring_strategy=scoring_strategy,
        )
        columns = first["columns"]
        items = first["items"]
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns)
        writer.writeheader()
        for row in items:
            writer.writerow(row)
        return output.getvalue()

    def export_personal_result_notice_xlsx(
        self,
        event_id: int,
        template_name: str,
        template_dir: str,
        layout_config_path: str,
    ) -> tuple[bytes, str]:
        safe_template_name = os.path.basename((template_name or "").strip())
        if not safe_template_name:
            raise ValueError("template_name 不能为空")
        if not safe_template_name.lower().endswith((".xlsx", ".xlsm")):
            raise ValueError("模板文件必须是 .xlsx 或 .xlsm")

        template_path = os.path.join(template_dir, safe_template_name)
        if not os.path.isfile(template_path):
            raise ValueError(f"模板文件不存在: {safe_template_name}")
        if not os.path.isfile(layout_config_path):
            raise ValueError("公示单坐标配置文件不存在")

        with open(layout_config_path, "r", encoding="utf-8") as f:
            layout = json.load(f)

        sheet_name = str(layout.get("sheet_name", "")).strip() or "Sheet1"
        environment_cells = layout.get("environment_cells", {}) or {}
        rank_rows = layout.get("rank_rows", []) or []
        if len(rank_rows) < 8:
            raise ValueError("坐标配置 rank_rows 至少需要 8 行")

        event, rows, env = self._get_personal_notice_payload(event_id)

        wb = load_workbook(template_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        env_values = {
            "date": env.get("date", ""),
            "wind_direction": env.get("wind_direction", ""),
            "wind_speed": env.get("wind_speed", ""),
            "air_quality": env.get("air_quality", ""),
            "weather": env.get("weather", ""),
            "temperature_high": env.get("temperature_high", ""),
            "temperature_low": env.get("temperature_low", ""),
            "event_name": self._event_display_name(event),
            "notice_title": self._notice_title_for_event(event, layout),
        }
        for key, cell in environment_cells.items():
            if key in env_values and cell:
                ws[str(cell)] = env_values[key]

        for idx in range(8):
            mapping = rank_rows[idx] if idx < len(rank_rows) else {}
            data = rows[idx] if idx < len(rows) else {}
            if mapping.get("rank"):
                ws[str(mapping["rank"])] = data.get("rank", idx + 1 if data else "")
            if mapping.get("name"):
                ws[str(mapping["name"])] = data.get("athlete_name", "")
            if mapping.get("department"):
                ws[str(mapping["department"])] = data.get("department_name", "")
            perf_cell = (
                mapping.get("performance")
                or mapping.get("perGormance")
                or mapping.get("成绩")
            )
            if perf_cell:
                ws[str(perf_cell)] = data.get("performance", "")

        buf = io.BytesIO()
        wb.save(buf)
        event_name_token = re.sub(r"[\\\\/:*?\"<>|]+", "_", self._event_display_name(event))
        filename = f"个人成绩公示单_{event_name_token}.xlsx"
        return buf.getvalue(), filename

    def _get_personal_notice_payload(self, event_id: int) -> tuple[dict, list[dict], dict[str, str]]:
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            event_row = repo.get_event_by_id(event_id)
            if not event_row:
                raise ValueError(f"项目不存在: {event_id}")
            event = dict(event_row)
            if int(event.get("is_individual", 0)) != 1:
                raise ValueError("仅个人项目支持导出个人成绩公示单")
            scoring_strategy = str(event.get("scoring_strategy", ""))
            rows = [dict(r) for r in repo.list_individual_results_for_event(event_id)]
            for row in rows:
                row["performance"] = self._format_performance_for_display(scoring_strategy, row.get("performance"))
        env = self.get_report_environment_settings()
        return event, rows, env

    def _convert_xlsx_to_pdf_with_excel(self, xlsx_path: str, pdf_path: str) -> None:
        try:
            import pythoncom  # type: ignore
            import win32com.client as win32  # type: ignore
        except Exception as exc:  # pragma: no cover
            raise RuntimeError("缺少 pywin32 或无法加载 win32com") from exc

        excel = None
        wb = None
        inited = False
        try:
            pythoncom.CoInitialize()
            inited = True
            excel = win32.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
            # 0 = PDF
            wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        finally:
            if wb is not None:
                try:
                    wb.Close(False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            if inited:
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass

    def _convert_xlsx_to_pdf_with_libreoffice(self, xlsx_path: str, out_dir: str) -> str:
        exe = shutil.which("soffice")
        if not exe:
            raise RuntimeError("未找到 soffice")
        cmd = [exe, "--headless", "--convert-to", "pdf", "--outdir", out_dir, xlsx_path]
        p = subprocess.run(cmd, capture_output=True, text=True)
        if p.returncode != 0:
            stderr = (p.stderr or "").strip()
            raise RuntimeError(f"soffice 转换失败: {stderr or p.returncode}")
        pdf_path = os.path.splitext(xlsx_path)[0] + ".pdf"
        if not os.path.exists(pdf_path):
            raise RuntimeError("soffice 未生成 pdf 文件")
        return pdf_path

    def _convert_xlsx_bytes_to_pdf_bytes(self, xlsx_bytes: bytes) -> bytes:
        errors: list[str] = []
        with tempfile.TemporaryDirectory(prefix="sports_notice_") as tmpdir:
            xlsx_path = os.path.join(tmpdir, "notice.xlsx")
            pdf_path = os.path.join(tmpdir, "notice.pdf")
            with open(xlsx_path, "wb") as f:
                f.write(xlsx_bytes)

            try:
                self._convert_xlsx_to_pdf_with_excel(xlsx_path, pdf_path)
                if not os.path.exists(pdf_path):
                    raise RuntimeError("Excel 未生成 pdf 文件")
                with open(pdf_path, "rb") as f:
                    return f.read()
            except Exception as exc:
                errors.append(f"Excel 转换失败: {exc}")

            try:
                generated_pdf = self._convert_xlsx_to_pdf_with_libreoffice(xlsx_path, tmpdir)
                with open(generated_pdf, "rb") as f:
                    return f.read()
            except Exception as exc:
                errors.append(f"LibreOffice 转换失败: {exc}")

        joined = "；".join(errors)
        raise ValueError(
            "xlsx 转 pdf 失败。请安装 Microsoft Excel（并安装 pywin32）或安装 LibreOffice(soffice)。"
            + (f" 详情：{joined}" if joined else "")
        )

    def export_personal_result_notice_pdf(
        self,
        event_id: int,
        template_name: str,
        template_dir: str,
        layout_config_path: str,
    ) -> tuple[bytes, str]:
        xlsx_bytes, xlsx_filename = self.export_personal_result_notice_xlsx(
            event_id=event_id,
            template_name=template_name,
            template_dir=template_dir,
            layout_config_path=layout_config_path,
        )
        pdf_bytes = self._convert_xlsx_bytes_to_pdf_bytes(xlsx_bytes)
        pdf_filename = re.sub(r"\.xlsx$", ".pdf", xlsx_filename, flags=re.IGNORECASE)
        return pdf_bytes, pdf_filename

    def get_initialization_status(self) -> dict:
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            meet_date_iso = repo.get_meet_date_iso()
            event_count = repo.events_count()
            department_count = repo.departments_count()
            athlete_count = repo.athletes_count()

            checks = [
                {
                    "key": "meet_date",
                    "label": "比赛日期已设置",
                    "ok": bool(meet_date_iso),
                    "detail": meet_date_iso or "未设置",
                },
                {
                    "key": "events",
                    "label": "项目数据已导入",
                    "ok": event_count > 0,
                    "detail": f"{event_count} 个项目",
                },
            ]

            completed = all(item["ok"] for item in checks)
            return {
                "completed": completed,
                "checks": checks,
                "summary": {
                    "event_count": event_count,
                    "department_count": department_count,
                    "athlete_count": athlete_count,
                },
            }

    def import_events_rows(self, rows: list[dict[str, str]]) -> dict:
        required = {"name", "category", "event_type", "gender", "age_group", "is_individual"}
        missing = required - set(rows[0].keys()) if rows else required
        if missing:
            raise ValueError(f"项目模板缺少字段: {sorted(missing)}")

        inserted = 0
        skipped = 0
        errors: list[str] = []
        skipped_details: list[str] = []

        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            for idx, row in enumerate(rows, start=2):
                try:
                    name = row["name"].strip()
                    category = row["category"].strip()
                    event_type = row["event_type"].strip()
                    scoring_strategy = (row.get("scoring_strategy") or "").strip()
                    gender_raw = row["gender"].strip()
                    age_group = row["age_group"].strip()
                    is_individual = int(row["is_individual"].strip())

                    if category not in {"competitive", "fun"}:
                        raise ValueError("category 必须为 competitive 或 fun")
                    if event_type not in {"track", "field", "fun"}:
                        raise ValueError("event_type 必须为 track/field/fun")
                    if scoring_strategy and scoring_strategy not in {"time", "length", "count"}:
                        raise ValueError("scoring_strategy 必须为 time/length/count")
                    target_genders = self._expand_event_genders(gender_raw)
                    if age_group not in {"A", "B", "C", "ALL"}:
                        raise ValueError("age_group 必须为 A/B/C/ALL")
                    if is_individual not in {0, 1}:
                        raise ValueError("is_individual 必须为 0 或 1")
                    if category == "fun" and event_type != "fun":
                        raise ValueError("趣味项目的 event_type 必须为 fun")
                    if category == "competitive" and event_type == "fun":
                        raise ValueError("竞技项目的 event_type 不能为 fun")
                    if (category == "fun" or is_individual == 0) and age_group != "ALL":
                        raise ValueError("趣味项目和集体项目必须使用 age_group=ALL")
                    if event_type in {"track", "field"}:
                        fixed = scoring_strategy_for_event_type(event_type)
                        if scoring_strategy and scoring_strategy != fixed:
                            raise ValueError(f"{event_type} 项目的 scoring_strategy 必须为 {fixed}")
                        scoring_strategy = fixed
                    else:
                        if not scoring_strategy:
                            scoring_strategy = "count"

                    for gender in target_genders:
                        if repo.event_exists(
                            name,
                            category,
                            event_type,
                            scoring_strategy,
                            gender,
                            age_group,
                            is_individual,
                        ):
                            skipped += 1
                            skipped_details.append(
                                f"第{idx}行: 已存在，已跳过（name={name}, gender={gender}, age_group={age_group}）"
                            )
                            continue

                        repo.insert_event(
                            name,
                            category,
                            event_type,
                            scoring_strategy,
                            gender,
                            age_group,
                            is_individual,
                        )
                        inserted += 1
                except Exception as exc:
                    errors.append(f"第{idx}行: {exc}")

            conn.commit()

        return {
            "inserted": inserted,
            "skipped": skipped,
            "errors": errors,
            "skipped_details": skipped_details,
        }

    def import_athletes_rows(self, rows: list[dict[str, str]], athlete_type: str) -> dict:
        athlete_type = self._validate_athlete_type(athlete_type)
        required = {"athlete_no", "name", "gender", "department_name"}
        missing = required - set(rows[0].keys()) if rows else required
        if missing:
            raise ValueError(f"名单模板缺少字段: {sorted(missing)}")

        inserted = 0
        skipped = 0
        errors: list[str] = []
        skipped_details: list[str] = []

        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            for idx, row in enumerate(rows, start=2):
                try:
                    athlete_no = (row.get("athlete_no") or "").strip()
                    if not athlete_no:
                        raise ValueError("athlete_no 不能为空")
                    name = (row.get("name") or "").strip()
                    gender = (row.get("gender") or "").strip()
                    department_name = (row.get("department_name") or "").strip()
                    if not name:
                        raise ValueError("name 不能为空")
                    if gender not in {"male", "female"}:
                        raise ValueError("gender 必须为 male/female")
                    if not department_name:
                        raise ValueError("department_name 不能为空")

                    age_group_raw = (row.get("age_group") or "").strip()
                    age_group = age_group_raw if age_group_raw else None
                    if age_group and age_group not in {"A", "B", "C"}:
                        raise ValueError("age_group 必须为 A/B/C")

                    total_members_text = (row.get("total_members") or "0").strip()
                    total_members = int(total_members_text) if total_members_text else 0

                    dept = repo.get_department_by_name(department_name)
                    if dept:
                        dept_id = int(dept["id"])
                        if total_members > int(dept["total_members"]):
                            repo.update_department_total_members(dept_id, total_members)
                    else:
                        dept_id = repo.insert_department(department_name, total_members)

                    exists = repo.get_athlete_by_no(athlete_type, athlete_no)
                    if exists:
                        skipped += 1
                        skipped_details.append(f"第{idx}行: athlete_no={athlete_no} 已存在，已跳过")
                        continue

                    repo.insert_athlete(
                        athlete_type=athlete_type,
                        athlete_no=athlete_no,
                        name=name,
                        gender=gender,
                        department_id=dept_id,
                        age_group=age_group,
                    )
                    inserted += 1
                except Exception as exc:
                    errors.append(f"第{idx}行: {exc}")
            conn.commit()

        return {
            "inserted": inserted,
            "skipped": skipped,
            "errors": errors,
            "skipped_details": skipped_details,
        }

    def import_registrations_rows(self, rows: list[dict[str, str]], target_category: str) -> dict:
        return self.import_registration_matrix_rows(rows, target_category)

    def _parse_event_ids_from_header(self, header: str) -> list[int]:
        text = (header or "").strip()
        matched = re.search(r"\[([0-9|,/\s]+)\]\s*$", text)
        if matched:
            raw = matched.group(1)
            parts = re.split(r"[|,/\s]+", raw)
            ids = [int(p) for p in parts if p and p.isdigit()]
            return ids
        lowered = text.lower()
        if lowered.startswith("event_id_"):
            suffix = lowered.replace("event_id_", "", 1).strip()
            if suffix.isdigit():
                return [int(suffix)]
        return []

    def _is_selected_mark(self, value: str) -> bool:
        mark = (value or "").strip().lower()
        return mark in {"1", "y", "yes", "true", "t", "x", "√", "是"}

    def import_registration_matrix_rows(self, rows: list[dict[str, str]], target_category: str) -> dict:
        if target_category not in {"competitive", "fun"}:
            raise ValueError("target_category 必须是 competitive 或 fun")
        required = {"athlete_no"}
        missing = required - set(rows[0].keys()) if rows else required
        if missing:
            raise ValueError(f"报名模板缺少字段: {sorted(missing)}")
        headers = list(rows[0].keys()) if rows else []
        event_columns: list[tuple[str, list[int]]] = []
        for header in headers:
            event_ids = self._parse_event_ids_from_header(header)
            if event_ids:
                event_columns.append((header, event_ids))
        if not event_columns:
            raise ValueError("未识别到项目列。请使用“项目名-组别[event_id]”或“项目名-组别[event_id1|event_id2]”。")

        inserted = 0
        skipped = 0
        errors: list[str] = []
        skipped_details: list[str] = []

        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            all_events = [dict(r) for r in repo.list_individual_events_by_category(target_category)]
            event_cache: dict[int, dict] = {int(e["id"]): e for e in all_events}
            for idx, row in enumerate(rows, start=2):
                try:
                    athlete_no = (row.get("athlete_no") or "").strip()
                    if not athlete_no:
                        raise ValueError("athlete_no 不能为空")

                    athlete = repo.get_athlete_by_no(target_category, athlete_no)
                    if not athlete:
                        raise ValueError(f"运动员不存在: {athlete_no}（请先导入名单）")
                    athlete_id = int(athlete["id"])

                    csv_gender = (row.get("gender") or "").strip()
                    if csv_gender and csv_gender != athlete["gender"]:
                        raise ValueError("gender 与已导入运动员信息不一致")
                    csv_group = (row.get("age_group") or "").strip()
                    athlete_group = athlete.get("age_group") or ""
                    if csv_group and csv_group != athlete_group:
                        raise ValueError("age_group 与已导入运动员信息不一致")

                    selected_any = False
                    for header, event_ids in event_columns:
                        if not self._is_selected_mark(str(row.get(header, ""))):
                            continue
                        selected_any = True

                        candidate_events: list[dict] = []
                        for event_id in event_ids:
                            event = event_cache.get(event_id, {})
                            if event:
                                candidate_events.append(event)
                        if not candidate_events:
                            errors.append(f"第{idx}行: 项目不存在: {event_ids}")
                            continue
                        chosen_event = None
                        for event in candidate_events:
                            if event["category"] == target_category and int(event["is_individual"]) == 1 and event["gender"] == athlete["gender"]:
                                chosen_event = event
                                break
                        if chosen_event is None:
                            for event in candidate_events:
                                if event["category"] == target_category and int(event["is_individual"]) == 1 and event["gender"] == "mixed":
                                    chosen_event = event
                                    break
                        if chosen_event is None:
                            errors.append(f"第{idx}行: 无法按性别自动匹配项目（候选ID={event_ids}）")
                            continue

                        try:
                            repo.insert_athlete_registration(target_category, athlete_id, int(chosen_event["id"]))
                            inserted += 1
                        except Exception:
                            skipped += 1
                            skipped_details.append(
                                f"第{idx}行: 重复报名或唯一约束冲突，已跳过（athlete_no={athlete_no}, event_id={chosen_event['id']}）"
                            )
                    if not selected_any:
                        skipped += 1
                        skipped_details.append(f"第{idx}行: 未勾选任何项目，已跳过（athlete_no={athlete_no}）")
                except Exception as exc:
                    errors.append(f"第{idx}行: {exc}")
            conn.commit()

        return {
            "inserted": inserted,
            "skipped": skipped,
            "errors": errors,
            "skipped_details": skipped_details,
        }

    def export_registration_template_csv(self, target_category: str, event_id: int) -> tuple[str, str]:
        # 兼容旧接口：按单项目导出两列表头
        if target_category not in {"competitive", "fun"}:
            raise ValueError("category 必须是 competitive 或 fun")
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            event = repo.get_event_by_id(event_id)
            if not event:
                raise ValueError(f"项目不存在: {event_id}")
            if event["category"] != target_category:
                raise ValueError(f"项目类别不匹配，应为 {target_category}")
            if int(event["is_individual"]) != 1:
                raise ValueError("仅个人项目支持报名模板导出")

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=["event_id", "athlete_no"])
        writer.writeheader()
        writer.writerow({"event_id": event_id, "athlete_no": ""})
        filename = f"{target_category}_registrations_event_{event_id}.csv"
        return output.getvalue(), filename

    def export_registration_matrix_template_csv(self, target_category: str) -> tuple[str, str]:
        if target_category not in {"competitive", "fun"}:
            raise ValueError("category 必须是 competitive 或 fun")

        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            events = [dict(row) for row in repo.list_individual_events_by_category(target_category)]
            athletes = [dict(row) for row in repo.list_athletes_by_type_with_department(target_category)]
            pairs = repo.list_registration_pairs_by_type(target_category)

        registered_map: dict[tuple[int, int], bool] = {}
        for row in pairs:
            registered_map[(int(row["athlete_ref_id"]), int(row["event_id"]))] = True

        grouped_events: list[tuple[str, list[dict]]] = []
        group_map: dict[str, list[dict]] = {}
        for event in events:
            key = f"{event['name']}|{event['age_group']}"
            if key not in group_map:
                group_map[key] = []
                grouped_events.append((key, group_map[key]))
            group_map[key].append(event)

        event_columns: list[str] = []
        for key, items in grouped_events:
            ref = items[0]
            ids = sorted(int(e["id"]) for e in items)
            ids_token = "|".join(str(i) for i in ids)
            event_columns.append(
                f"{ref['name']}-{self._event_group_label(str(ref['age_group']))}[{ids_token}]"
            )
        fieldnames = ["athlete_no", "name", "gender", "age_group", "department_name"] + event_columns
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()

        for athlete in athletes:
            row = {
                "athlete_no": athlete.get("athlete_no", "") or "",
                "name": athlete.get("name", "") or "",
                "gender": athlete.get("gender", "") or "",
                "age_group": athlete.get("age_group", "") or "",
                "department_name": athlete.get("department_name", "") or "",
            }
            athlete_id = int(athlete["athlete_ref_id"])
            for (key, items), col in zip(grouped_events, event_columns):
                marked = any(registered_map.get((athlete_id, int(e["id"]))) for e in items)
                row[col] = "1" if marked else ""
            writer.writerow(row)

        filename = f"{target_category}_registrations_matrix.csv"
        return output.getvalue(), filename

    def clear_table_data(
        self,
        requested_tables: list[str],
        confirm_text: str,
        confirm_code: str,
        acknowledged: bool,
    ) -> dict:
        selected = sorted({t for t in requested_tables if t in self.CLEAR_TABLES})
        if not selected:
            raise ValueError("请至少选择一张表")
        if not acknowledged:
            raise ValueError("请先勾选确认选项")
        if (confirm_text or "").strip().upper() != "DELETE":
            raise ValueError("请正确输入确认口令 DELETE")
        expected_code = f"CLEAR-{len(selected)}"
        if (confirm_code or "").strip().upper() != expected_code:
            raise ValueError(f"校验码错误，应为 {expected_code}")

        counts: dict[str, int] = {k: 0 for k in self.CLEAR_TABLES}

        def _exec_delete(conn, table_key: str, sql: str, params: tuple = ()) -> None:
            cur = conn.execute(sql, params)
            delta = cur.rowcount if cur.rowcount and cur.rowcount > 0 else 0
            counts[table_key] += delta

        with self.db.connect() as conn:
            clear_settings = "settings" in selected
            clear_departments = "departments" in selected
            clear_events = "events" in selected
            clear_comp = "competitive_athletes" in selected or clear_departments
            clear_fun = "fun_athletes" in selected or clear_departments
            clear_teams = "teams" in selected or clear_events or clear_departments
            clear_team_members = "team_members" in selected
            clear_regs = "athlete_registrations" in selected
            clear_results = "results" in selected

            if clear_departments:
                _exec_delete(conn, "results", "DELETE FROM results")
                _exec_delete(conn, "athlete_registrations", "DELETE FROM athlete_registrations")
                _exec_delete(conn, "team_members", "DELETE FROM team_members")
                _exec_delete(conn, "teams", "DELETE FROM teams")
                _exec_delete(conn, "competitive_athletes", "DELETE FROM competitive_athletes")
                _exec_delete(conn, "fun_athletes", "DELETE FROM fun_athletes")
                _exec_delete(conn, "departments", "DELETE FROM departments")
            else:
                if clear_events:
                    _exec_delete(conn, "results", "DELETE FROM results")
                    _exec_delete(conn, "athlete_registrations", "DELETE FROM athlete_registrations")
                    _exec_delete(conn, "team_members", "DELETE FROM team_members")
                    _exec_delete(conn, "teams", "DELETE FROM teams")
                    _exec_delete(conn, "events", "DELETE FROM events")
                else:
                    if clear_teams:
                        _exec_delete(conn, "results", "DELETE FROM results WHERE team_id IS NOT NULL")
                        _exec_delete(conn, "team_members", "DELETE FROM team_members")
                        _exec_delete(conn, "teams", "DELETE FROM teams")
                    elif clear_team_members:
                        _exec_delete(conn, "team_members", "DELETE FROM team_members")

                    if clear_comp:
                        _exec_delete(conn, "results", "DELETE FROM results WHERE athlete_type='competitive'")
                        _exec_delete(
                            conn,
                            "athlete_registrations",
                            "DELETE FROM athlete_registrations WHERE athlete_type='competitive'",
                        )
                        _exec_delete(conn, "team_members", "DELETE FROM team_members WHERE athlete_type='competitive'")
                        _exec_delete(conn, "competitive_athletes", "DELETE FROM competitive_athletes")

                    if clear_fun:
                        _exec_delete(conn, "results", "DELETE FROM results WHERE athlete_type='fun'")
                        _exec_delete(
                            conn,
                            "athlete_registrations",
                            "DELETE FROM athlete_registrations WHERE athlete_type='fun'",
                        )
                        _exec_delete(conn, "team_members", "DELETE FROM team_members WHERE athlete_type='fun'")
                        _exec_delete(conn, "fun_athletes", "DELETE FROM fun_athletes")

                    if clear_results:
                        _exec_delete(conn, "results", "DELETE FROM results")
                    if clear_regs:
                        _exec_delete(conn, "athlete_registrations", "DELETE FROM athlete_registrations")
                    if clear_team_members and not clear_teams:
                        _exec_delete(conn, "team_members", "DELETE FROM team_members")

            if clear_settings:
                _exec_delete(conn, "settings", "DELETE FROM settings")

            conn.commit()

        affected = {k: v for k, v in counts.items() if v > 0}
        return {
            "requested_tables": selected,
            "expected_code": expected_code,
            "deleted_rows": affected,
        }
