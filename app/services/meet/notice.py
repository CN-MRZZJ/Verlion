import io
import json
import os
import re
import shutil
import subprocess
import tempfile

from openpyxl import load_workbook

from app.models import SportsRepository


class MeetNoticeMixin:
    def export_personal_result_notice_xlsx(
        self,
        event_id: int,
        template_name: str,
        template_dir: str,
        layout_config_path: str,
    ) -> tuple[bytes, str]:
        safe_template_name = os.path.basename((template_name or "").strip())
        if not safe_template_name:
            raise ValueError("template_name 不能为空")
        if not safe_template_name.lower().endswith((".xlsx", ".xlsm")):
            raise ValueError("模板文件必须是 .xlsx 或 .xlsm")

        template_path = os.path.join(template_dir, safe_template_name)
        if not os.path.isfile(template_path):
            raise ValueError(f"模板文件不存在: {safe_template_name}")
        if not os.path.isfile(layout_config_path):
            raise ValueError("公示单坐标配置文件不存在")

        with open(layout_config_path, "r", encoding="utf-8") as f:
            layout = json.load(f)

        sheet_name = str(layout.get("sheet_name", "")).strip() or "Sheet1"
        environment_cells = layout.get("environment_cells", {}) or {}
        rank_rows = layout.get("rank_rows", []) or []
        if len(rank_rows) < 8:
            raise ValueError("坐标配置 rank_rows 至少需要 8 行")

        event, rows, env = self._get_personal_notice_payload(event_id)

        wb = load_workbook(template_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        env_values = {
            "date": env.get("date", ""),
            "wind_direction": env.get("wind_direction", ""),
            "wind_speed": env.get("wind_speed", ""),
            "air_quality": env.get("air_quality", ""),
            "weather": env.get("weather", ""),
            "temperature_high": env.get("temperature_high", ""),
            "temperature_low": env.get("temperature_low", ""),
            "event_name": self._event_display_name(event),
            "notice_title": self._notice_title_for_event(event, layout),
        }
        for key, cell in environment_cells.items():
            if key in env_values and cell:
                ws[str(cell)] = env_values[key]

        for idx in range(8):
            mapping = rank_rows[idx] if idx < len(rank_rows) else {}
            data = rows[idx] if idx < len(rows) else {}
            if mapping.get("rank"):
                ws[str(mapping["rank"])] = data.get("rank", idx + 1 if data else "")
            if mapping.get("name"):
                ws[str(mapping["name"])] = data.get("athlete_name", "")
            if mapping.get("department"):
                ws[str(mapping["department"])] = data.get("department_name", "")
            perf_cell = mapping.get("performance") or mapping.get("perGormance") or mapping.get("成绩")
            if perf_cell:
                ws[str(perf_cell)] = data.get("performance", "")

        buf = io.BytesIO()
        wb.save(buf)
        event_name_token = re.sub(r"[\\/:*?\"<>|]+", "_", self._event_display_name(event))
        filename = f"个人成绩公示单_{event_name_token}.xlsx"
        return buf.getvalue(), filename

    def _get_personal_notice_payload(self, event_id: int) -> tuple[dict, list[dict], dict[str, str]]:
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            event_row = repo.get_event_by_id(event_id)
            if not event_row:
                raise ValueError(f"项目不存在: {event_id}")
            event = dict(event_row)
            if int(event.get("is_individual", 0)) != 1:
                raise ValueError("仅个人项目支持导出个人成绩公示单")
            scoring_strategy = str(event.get("scoring_strategy", ""))
            rows = [dict(r) for r in repo.list_individual_results_for_event(event_id)]
            for row in rows:
                row["performance"] = self._format_performance_for_display(scoring_strategy, row.get("performance"))
        env = self.get_report_environment_settings()
        return event, rows, env

    def _convert_xlsx_to_pdf_with_excel(self, xlsx_path: str, pdf_path: str) -> None:
        try:
            import pythoncom  # type: ignore
            import win32com.client as win32  # type: ignore
        except Exception as exc:  # pragma: no cover
            raise RuntimeError("缺少 pywin32 或无法加载 win32com") from exc

        excel = None
        wb = None
        inited = False
        try:
            pythoncom.CoInitialize()
            inited = True
            excel = win32.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
            wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        finally:
            if wb is not None:
                try:
                    wb.Close(False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            if inited:
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass

    def _convert_xlsx_to_pdf_with_libreoffice(self, xlsx_path: str, out_dir: str) -> str:
        exe = shutil.which("soffice")
        if not exe:
            raise RuntimeError("未找到 soffice")
        cmd = [exe, "--headless", "--convert-to", "pdf", "--outdir", out_dir, xlsx_path]
        p = subprocess.run(cmd, capture_output=True, text=True)
        if p.returncode != 0:
            stderr = (p.stderr or "").strip()
            raise RuntimeError(f"soffice 转换失败: {stderr or p.returncode}")
        pdf_path = os.path.splitext(xlsx_path)[0] + ".pdf"
        if not os.path.exists(pdf_path):
            raise RuntimeError("soffice 未生成 pdf 文件")
        return pdf_path

    def _convert_xlsx_bytes_to_pdf_bytes(self, xlsx_bytes: bytes) -> bytes:
        errors: list[str] = []
        with tempfile.TemporaryDirectory(prefix="sports_notice_") as tmpdir:
            xlsx_path = os.path.join(tmpdir, "notice.xlsx")
            pdf_path = os.path.join(tmpdir, "notice.pdf")
            with open(xlsx_path, "wb") as f:
                f.write(xlsx_bytes)

            try:
                self._convert_xlsx_to_pdf_with_excel(xlsx_path, pdf_path)
                if not os.path.exists(pdf_path):
                    raise RuntimeError("Excel 未生成 pdf 文件")
                with open(pdf_path, "rb") as f:
                    return f.read()
            except Exception as exc:
                errors.append(f"Excel 转换失败: {exc}")

            try:
                generated_pdf = self._convert_xlsx_to_pdf_with_libreoffice(xlsx_path, tmpdir)
                with open(generated_pdf, "rb") as f:
                    return f.read()
            except Exception as exc:
                errors.append(f"LibreOffice 转换失败: {exc}")

        joined = "；".join(errors)
        raise ValueError(
            "xlsx 转 pdf 失败。请安装 Microsoft Excel（并安装 pywin32）或安装 LibreOffice(soffice)。"
            + (f" 详情：{joined}" if joined else "")
        )

    def export_personal_result_notice_pdf(
        self,
        event_id: int,
        template_name: str,
        template_dir: str,
        layout_config_path: str,
    ) -> tuple[bytes, str]:
        xlsx_bytes, xlsx_filename = self.export_personal_result_notice_xlsx(
            event_id=event_id,
            template_name=template_name,
            template_dir=template_dir,
            layout_config_path=layout_config_path,
        )
        pdf_bytes = self._convert_xlsx_bytes_to_pdf_bytes(xlsx_bytes)
        pdf_filename = re.sub(r"\.xlsx$", ".pdf", xlsx_filename, flags=re.IGNORECASE)
        return pdf_bytes, pdf_filename
