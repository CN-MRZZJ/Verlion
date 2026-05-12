import io
import json
import os
import re

from openpyxl import load_workbook

from app.models.repositories import SportsRepository

_ROUND_NAMES_BY_COUNT = {
    1: {1: "决赛"},
    2: {1: "预赛", 2: "决赛"},
    3: {1: "预赛", 2: "半决赛", 3: "决赛"},
    4: {1: "预赛", 2: "复赛", 3: "半决赛", 4: "决赛"},
}


class MeetNoticeMixin:
    def _round_name_for_notice(self, event_id: int, round_id: int) -> str:
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            config = repo.get_heats_config(event_id)
            heat_rounds = int(config["heat_rounds"]) if config else 1
        names = _ROUND_NAMES_BY_COUNT.get(heat_rounds, _ROUND_NAMES_BY_COUNT[1])
        return names.get(round_id, f"第{round_id}轮")

    def _max_notice_rows(self, repo: SportsRepository, event: dict, round_id: int) -> int:
        config = repo.get_heats_config(int(event["id"]))
        heat_rounds = int(config["heat_rounds"]) if config else 1
        if round_id < heat_rounds:
            return 9999
        is_individual = int(event.get("is_individual", 1))
        result_type = "individual" if is_individual else "team"
        rows = repo.conn.execute(
            "SELECT MAX(rank) AS mx FROM point_rules WHERE result_type=?", (result_type,)
        ).fetchone()
        return int(rows["mx"]) if rows and rows["mx"] else 8

    def _export_result_notice_xlsx(
        self, event_id: int, round_id: int, template_name: str,
        layout_config_path: str, group_by_heat: bool,
    ) -> tuple[bytes, str]:
        safe_name = os.path.basename((template_name or "").strip())
        if not safe_name or not safe_name.lower().endswith((".xlsx", ".xlsm")):
            raise ValueError("模板必须是 .xlsx 或 .xlsm")

        template_dir = os.path.dirname(layout_config_path)
        template_path = os.path.join(template_dir, safe_name)
        if not os.path.isfile(template_path):
            raise ValueError(f"模板文件不存在: {safe_name}")
        if not os.path.isfile(layout_config_path):
            raise ValueError("布局配置文件不存在")

        with open(layout_config_path, "r", encoding="utf-8") as f:
            layout = json.load(f)

        sheet_name = str(layout.get("sheet_name", "")).strip() or "Sheet1"
        environment_cells = layout.get("environment_cells", {}) or {}
        row_template = layout.get("row_template", {}) or {}
        start_row = int(layout.get("start_row", 15))

        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            event_row = repo.get_event_by_id(event_id)
            if not event_row:
                raise ValueError(f"项目不存在: {event_id}")
            event = dict(event_row)
            scoring_strategy = str(event.get("scoring_strategy", ""))
            is_individual = int(event.get("is_individual", 1)) == 1

            if is_individual:
                raw_rows = [dict(r) for r in repo.list_individual_results_for_event_all(event_id, round_id=round_id)]
            else:
                raw_rows = [dict(r) for r in repo.list_team_results_for_event_all(event_id, round_id=round_id)]

            # Sort by overall rank
            def _sort_key(d):
                rk = d.get("rank", 999)
                return int(rk) if rk is not None else 999
            ranked = sorted(raw_rows, key=_sort_key)

            # heat_name lookup
            grouped = [dict(r) for r in repo.list_results_grouped_by_heat(event_id, round_id)]
            heat_lookup = {g["result_id"]: g.get("heat_name", "") for g in grouped}

            for row in ranked:
                row["performance"] = self._format_performance_for_display(
                    scoring_strategy, row.get("performance"))
                row["name"] = row.get("athlete_name") or row.get("team_name", "")
                row["department"] = row.get("department_name", "")
                row["heat_name"] = heat_lookup.get(int(row["id"]), "")
                row["heat_rank"] = row.get("heat_rank", "")

            max_rows = self._max_notice_rows(repo, event, round_id)

            if group_by_heat:
                heat_groups: dict[str, list[dict]] = {}
                heat_order: list[str] = []
                for row in ranked:
                    hn = row.get("heat_name", "")
                    if hn not in heat_groups:
                        heat_groups[hn] = []
                        heat_order.append(hn)
                    if len(heat_groups[hn]) < max_rows:
                        heat_groups[hn].append(row)
            else:
                rows = ranked[:max_rows]

        round_name = self._round_name_for_notice(event_id, round_id)
        env_values = {
            "date": self._get_notice_env("date"),
            "weather": self._get_notice_env("weather"),
            "wind_direction": self._get_notice_env("wind_direction"),
            "wind_speed": self._get_notice_env("wind_speed"),
            "air_quality": self._get_notice_env("air_quality"),
            "temperature_high": self._get_notice_env("temperature_high"),
            "temperature_low": self._get_notice_env("temperature_low"),
            "event_name": self._event_display_name(event),
            "round_name": round_name,
            "notice_title": self._notice_title_for_event(event, layout),
        }

        wb = load_workbook(template_path)
        template_ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        def _fill_sheet(ws, rows_to_fill, heat_name: str = ""):
            ev = dict(env_values)
            if heat_name:
                ev["event_name"] = f"{ev['event_name']} - {heat_name}"
            # Build merge anchor map
            merge_anchor: dict[str, str] = {}
            for mr in ws.merged_cells.ranges:
                anchor = str(mr).split(":")[0]
                from openpyxl.utils import range_boundaries, get_column_letter
                min_col, min_row, max_col, max_row = range_boundaries(str(mr))
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        merge_anchor[f"{get_column_letter(c)}{r}"] = anchor

            def _cell(col: str, row_num: int) -> str:
                addr = f"{col}{row_num}"
                return merge_anchor.get(addr, addr)

            for key, cell in environment_cells.items():
                if key in ev and cell:
                    ws[str(cell)] = ev[key]

            for idx, row in enumerate(rows_to_fill):
                row_num = start_row + idx
                for key, col in row_template.items():
                    if col:
                        addr = _cell(col, row_num)
                        val = row.get(key, "")
                        ws[addr] = val if val is not None else ""

        event_token = re.sub(r"[\\/:*?\"<>|]+", "_", self._event_display_name(event))

        if group_by_heat:
            # One xlsx per heat, zipped together — no copy_worksheet
            import zipfile
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for hn in heat_order:
                    wb_heat = load_workbook(template_path)
                    ws_heat = wb_heat[sheet_name] if sheet_name in wb_heat.sheetnames else wb_heat.active
                    ws_heat.title = hn
                    _fill_sheet(ws_heat, heat_groups[hn], hn)
                    heat_buf = io.BytesIO()
                    wb_heat.save(heat_buf)
                    safe_hn = re.sub(r"[\\/:*?\"<>|]+", "_", hn)
                    zf.writestr(f"{safe_hn}.xlsx", heat_buf.getvalue())
            filename = f"分组公告_{event_token}_{round_name}.zip"
            return zip_buf.getvalue(), filename
        else:
            _fill_sheet(template_ws, rows)
            buf = io.BytesIO()
            wb.save(buf)
            filename = f"全部公告_{event_token}_{round_name}.xlsx"
            return buf.getvalue(), filename

    def _get_notice_env(self, key: str) -> str:
        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            val = repo.get_setting(f"report_env.{key}")
            return str(val) if val else ""

    def export_grouped_result_notice_xlsx(
        self, event_id: int, round_id: int,
        template_name: str, layout_config_path: str,
    ) -> tuple[bytes, str]:
        return self._export_result_notice_xlsx(
            event_id, round_id, template_name, layout_config_path, group_by_heat=True)

    def export_full_result_notice_xlsx(
        self, event_id: int, round_id: int,
        template_name: str, layout_config_path: str,
    ) -> tuple[bytes, str]:
        return self._export_result_notice_xlsx(
            event_id, round_id, template_name, layout_config_path, group_by_heat=False)

    def export_grouped_result_notice_pdf(
        self, event_id: int, round_id: int,
        template_name: str, layout_config_path: str,
    ) -> tuple[bytes, str]:
        import zipfile
        xlsx_bytes, xlsx_name = self.export_grouped_result_notice_xlsx(
            event_id, round_id, template_name, layout_config_path)
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf_out:
            with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as zf_in:
                for name in zf_in.namelist():
                    pdf = self._convert_xlsx_bytes_to_pdf_bytes(zf_in.read(name))
                    zf_out.writestr(name.replace(".xlsx", ".pdf"), pdf)
        pdf_name = xlsx_name.replace(".zip", ".pdf.zip") if xlsx_name.endswith(".zip") else xlsx_name + ".pdf.zip"
        return zip_buf.getvalue(), pdf_name

    def export_full_result_notice_pdf(
        self, event_id: int, round_id: int,
        template_name: str, layout_config_path: str,
    ) -> tuple[bytes, str]:
        xlsx_bytes, xlsx_name = self.export_full_result_notice_xlsx(
            event_id, round_id, template_name, layout_config_path)
        pdf_bytes = self._convert_xlsx_bytes_to_pdf_bytes(xlsx_bytes)
        return pdf_bytes, re.sub(r"\.xlsx$", ".pdf", xlsx_name, flags=re.IGNORECASE)

    def _convert_xlsx_bytes_to_pdf_bytes(self, xlsx_bytes: bytes) -> bytes:
        import shutil
        import subprocess
        import tempfile

        errors: list[str] = []
        with tempfile.TemporaryDirectory(prefix="sports_notice_") as tmpdir:
            xlsx_path = os.path.join(tmpdir, "notice.xlsx")
            pdf_path = os.path.join(tmpdir, "notice.pdf")
            with open(xlsx_path, "wb") as f:
                f.write(xlsx_bytes)

            try:
                import pythoncom
                import win32com.client as win32
                pythoncom.CoInitialize()
                try:
                    excel = win32.DispatchEx("Excel.Application")
                    excel.Visible = False
                    excel.DisplayAlerts = False
                    wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
                    wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
                    wb.Close(False)
                    excel.Quit()
                finally:
                    pythoncom.CoUninitialize()
                if os.path.exists(pdf_path):
                    with open(pdf_path, "rb") as f:
                        return f.read()
            except Exception as exc:
                errors.append(f"Excel 转换失败: {exc}")

            exe = shutil.which("soffice")
            if exe:
                try:
                    p = subprocess.run(
                        [exe, "--headless", "--convert-to", "pdf", "--outdir", tmpdir, xlsx_path],
                        capture_output=True, text=True,
                    )
                    if p.returncode == 0 and os.path.exists(pdf_path):
                        with open(pdf_path, "rb") as f:
                            return f.read()
                    errors.append(f"soffice 失败: {p.stderr}")
                except Exception as exc:
                    errors.append(f"LibreOffice 转换失败: {exc}")

        raise ValueError("xlsx 转 pdf 失败。请安装 Excel + pywin32 或 LibreOffice。" + "；".join(errors))
