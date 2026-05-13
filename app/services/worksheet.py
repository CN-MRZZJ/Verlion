import io
import json
import os
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from app.models.repositories import SportsRepository


class MeetWorksheetMixin:
    def export_checkin_xlsx(
        self, event_id: int, round_id: int,
        template_name: str, layout_config_path: str,
        heat_id: int | None = None,
        field_layout_path: str | None = None,
    ) -> tuple[bytes, str]:
        safe_name = os.path.basename((template_name or "").strip())
        if not safe_name or not safe_name.lower().endswith((".xlsx", ".xlsm")):
            raise ValueError("模板必须是 .xlsx 或 .xlsm")

        template_dir = os.path.dirname(layout_config_path)
        template_path = os.path.join(template_dir, safe_name)
        if not os.path.isfile(template_path):
            raise ValueError(f"模板文件不存在: {safe_name}")
        if not os.path.isfile(layout_config_path):
            raise ValueError("布局配置文件不存在")

        with open(layout_config_path, "r", encoding="utf-8") as f:
            layout = json.load(f)

        sheet_name = str(layout.get("sheet_name", "")).strip() or "Sheet1"
        environment_cells = layout.get("environment_cells", {}) or {}
        heat_block = layout.get("heat_block", {}) or {}

        with self.db.connect() as conn:
            repo = SportsRepository(conn)
            event_row = repo.get_event_by_id(event_id)
            if not event_row:
                raise ValueError(f"项目不存在: {event_id}")
            event = dict(event_row)
            event_type = str(event.get("event_type", ""))

            if event_type != "track" and field_layout_path and os.path.isfile(field_layout_path):
                layout_config_path = field_layout_path
                with open(layout_config_path, "r", encoding="utf-8") as f2:
                    layout = json.load(f2)
                sheet_name = str(layout.get("sheet_name", "")).strip() or "Sheet1"
                environment_cells = layout.get("environment_cells", {}) or {}
                heat_block = layout.get("heat_block", {}) or {}

            heats_data: list[dict] = []
            rounds = repo.list_rounds(event_id)
            for r in rounds:
                rd = dict(r)
                if int(rd["round_number"]) != round_id:
                    continue
                for h in repo.list_heats(rd["id"]):
                    ht = dict(h)
                    if heat_id is not None and int(ht["id"]) != heat_id:
                        continue
                    entries = sorted(
                        [dict(e) for e in repo.list_heat_entries(ht["id"])],
                        key=lambda x: x["lane"] or 999,
                    )
                    heats_data.append({"heat_name": ht["heat_name"], "entries": entries})

            if not heats_data:
                raise ValueError("没有找到编排数据，请先生成编排")

        round_name = self._round_name_for_notice(event_id, round_id)
        event_display = self._event_display_name(event)
        env_values = {
            "date": self._get_notice_env("date"),
            "weather": self._get_notice_env("weather"),
            "wind_direction": self._get_notice_env("wind_direction"),
            "wind_speed": self._get_notice_env("wind_speed"),
            "air_quality": self._get_notice_env("air_quality"),
            "temperature_high": self._get_notice_env("temperature_high"),
            "temperature_low": self._get_notice_env("temperature_low"),
            "event_name": event_display,
            "round_name": round_name,
        }

        is_track = event_type == "track"
        anchor_row = int(heat_block.get("anchor_row", 1))
        block_height = int(heat_block.get("block_height", 8))
        start_col_letter = str(heat_block.get("start_col", "B"))
        start_col = ord(start_col_letter) - ord("A") + 1
        lanes = int(heat_block.get("lanes", 8))
        rows_cfg = heat_block.get("rows", {})
        row_template = layout.get("row_template", {}) or {}
        start_row = int(layout.get("start_row", 6))
        max_rows = int(layout.get("max_rows", 30))

        wb = load_workbook(template_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        # Clone the template page for additional heats
        if len(heats_data) > 1:
            _clone_page_block(ws, anchor_row, block_height, len(heats_data))

        _fill_sheet(ws, heats_data, env_values, environment_cells,
                    anchor_row, block_height, start_col, lanes,
                    heat_block.get("info_row", 3), rows_cfg, row_template,
                    start_row, max_rows, layout, is_track,
                    heat_block.get("heat_name_col", ""))

        buf = io.BytesIO()
        wb.save(buf)
        event_token = re.sub(r"[\\/:*?\"<>|]+", "_", event_display)
        if heat_id is not None:
            hd = heats_data[0]
            filename = f"检录单_{event_token}_{round_name}_{hd['heat_name']}.xlsx"
        else:
            filename = f"检录单_{event_token}_{round_name}.xlsx"
        return buf.getvalue(), filename


def _clone_page_block(ws, anchor_row: int, block_height: int, total_blocks: int):
    """Clone the template page downward for additional heat blocks,
    preserving merged cells, row heights, cell styles and values."""
    # Save original state
    orig_merges = []
    for mr in ws.merged_cells.ranges:
        parts = str(mr).split(":")
        if len(parts) == 2:
            orig_merges.append((str(mr), parts[0], parts[1]))
    row_heights = {}
    for r in range(anchor_row, anchor_row + block_height):
        h = ws.row_dimensions[r].height
        if h:
            row_heights[r] = h

    # Unmerge all
    for _, _, _ in orig_merges:
        for mr in list(ws.merged_cells.ranges):
            try:
                ws.unmerge_cells(str(mr))
            except Exception:
                pass

    # Insert rows and copy cell content for each extra block
    for block_n in range(1, total_blocks):
        target_start = anchor_row + block_n * block_height
        ws.insert_rows(target_start, block_height)
        for dr in range(block_height):
            src_row = anchor_row + dr
            tgt_row = target_start + dr
            if src_row in row_heights:
                ws.row_dimensions[tgt_row].height = row_heights[src_row]
            for col in range(1, ws.max_column + 1):
                src = ws.cell(row=src_row, column=col)
                tgt = ws.cell(row=tgt_row, column=col)
                if src.value is not None:
                    tgt.value = src.value
                if src.has_style:
                    tgt.font = src.font.copy()
                    tgt.border = src.border.copy()
                    tgt.fill = src.fill.copy()
                    tgt.number_format = src.number_format
                    tgt.alignment = src.alignment.copy()

    # Restore merged cells for all blocks
    for block_n in range(total_blocks):
        offset = block_n * block_height
        for _, tl, br in orig_merges:
            m1 = re.match(r'([A-Z]+)(\d+)', tl)
            m2 = re.match(r'([A-Z]+)(\d+)', br)
            if m1 and m2:
                tl_new = f"{m1.group(1)}{int(m1.group(2)) + offset}"
                br_new = f"{m2.group(1)}{int(m2.group(2)) + offset}"
                try:
                    ws.merge_cells(f"{tl_new}:{br_new}")
                except Exception:
                    pass


def _fill_sheet(ws, heat_list, env_values, environment_cells,
                anchor_row, block_height, start_col, lanes,
                info_row, rows_cfg, row_template, start_row,
                max_rows, layout, is_track, heat_name_col):
    """Write environment info and heat data into the sheet."""

    merge_anchor: dict[str, str] = {}
    for mr in ws.merged_cells.ranges:
        anchor = str(mr).split(":")[0]
        min_col, min_row, max_col, max_row = range_boundaries(str(mr))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merge_anchor[f"{get_column_letter(c)}{r}"] = anchor

    def _safe_ws(addr: str, value):
        target = merge_anchor.get(addr, addr)
        ws[target] = value if value is not None else ""

    for key, cell in environment_cells.items():
        if key in env_values and cell:
            _safe_ws(str(cell), env_values[key])

    for heat_idx, hd in enumerate(heat_list):
        base_row = anchor_row + heat_idx * block_height
        hn = hd["heat_name"]

        # Info row
        ir = base_row + int(info_row) - anchor_row
        _safe_ws(f"A{ir}", env_values.get("event_name", ""))
        if is_track:
            _safe_ws(f"E{ir}", env_values.get("round_name", ""))
        else:
            _safe_ws(f"H{ir}", env_values.get("round_name", ""))
        if heat_name_col:
            _safe_ws(f"{heat_name_col}{ir}", hn)
        else:
            heat_col = get_column_letter(start_col + lanes - 1)
            _safe_ws(f"{heat_col}{ir}", hn)

        entries = sorted(hd["entries"], key=lambda e: e["lane"] or 999)

        if is_track:
            # Horizontal layout: lane columns
            if "lane_labels" in rows_cfg:
                lr = base_row + int(rows_cfg["lane_labels"])
                labels = layout.get("lane_labels", [f"{l}道" for l in range(1, lanes + 1)])
                for l in range(lanes):
                    _safe_ws(f"{get_column_letter(start_col + l)}{lr}",
                             labels[l] if l < len(labels) else f"{l + 1}道")

            lane_map: dict[int, dict] = {}
            for e in entries:
                ln = int(e["lane"]) if e["lane"] is not None else -1
                if ln > 0:
                    lane_map[ln] = e

            for field_name, field_offset in rows_cfg.items():
                if field_name in ("header", "lane_labels", "info_row"):
                    continue
                field_offset = int(field_offset)
                for l in range(1, lanes + 1):
                    col = get_column_letter(start_col + l - 1)
                    row = base_row + field_offset
                    entry = lane_map.get(l, {})
                    _safe_ws(f"{col}{row}", entry.get(field_name, ""))
        else:
            # Vertical layout: one row per athlete
            for idx, entry in enumerate(entries[:max_rows]):
                row_num = base_row + start_row - anchor_row + idx
                for key, col in row_template.items():
                    if col:
                        _safe_ws(f"{col}{row_num}", entry.get(key, ""))
